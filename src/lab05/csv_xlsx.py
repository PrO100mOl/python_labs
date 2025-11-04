from pathlib import Path
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def csv_to_xlsx(csv_path: str | Path, xlsx_path: str | Path) -> None:
    p_csv = Path(csv_path)
    p_xlsx = Path(xlsx_path)

    if p_csv.suffix.lower() != ".csv" or p_xlsx.suffix.lower() != ".xlsx":
        raise ValueError("Неверный тип файла: ожидаются .csv (вход) и .xlsx (выход).")

    with p_csv.open(encoding="utf-8") as f:
        sample = f.read(4096)
        if not sample.strip():
            raise ValueError("Пустой CSV.")
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        except csv.Error:
            dialect = csv.get_dialect("excel")
        reader = csv.reader(f, dialect=dialect)
        rows = list(reader)

    if not rows:
        raise ValueError("CSV пуст.")
    header = rows[0]
    if not header or all((h or "").strip() == "" for h in header):
        raise ValueError("CSV без заголовка.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for row in rows:
        ws.append(row)

    # Автоширина колонок
    col_widths = [0] * len(header)
    for row in rows:
        for i, cell in enumerate(row):
            length = len(str(cell)) if cell is not None else 0
            if i >= len(col_widths):
                col_widths.extend([0] * (i + 1 - len(col_widths)))
            if length > col_widths[i]:
                col_widths[i] = length

    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(8, width)

    wb.save(p_xlsx)

csv_to_xlsx('data/samples/people.csv','data/out/people.xlsx')